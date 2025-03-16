from collections import defaultdict
from io import BytesIO
from typing import Dict, List, Tuple, Union

import openpyxl
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from starlette.responses import StreamingResponse

from database.models import ActivityRegistry, NormMaterials, Norms, User
from database.models.material_types import MaterialTypes


async def update_material_type_db(
    db_session: AsyncSession, material_type: MaterialTypes, **update_data: dict
) -> MaterialTypes:
    for field, value in update_data.items():
        setattr(material_type, field, value)
    await db_session.commit()
    await db_session.refresh(material_type)
    return material_type


async def _get_norm_materials(
    db_session: AsyncSession,
    material_id: int,
) -> List['NormMaterials']:
    query = select(NormMaterials).where(
        NormMaterials.material_type_id == material_id
    )
    result = await db_session.scalars(query)
    return result.all()


async def _get_norms_for_materials(
    db_session: AsyncSession,
    norm_materials: List['NormMaterials'],
) -> Dict[int, 'Norms']:
    norm_ids = {nm.norm_id for nm in norm_materials}
    if not norm_ids:
        return {}
    query = select(Norms).where(Norms.id.in_(norm_ids))
    norms = await db_session.scalars(query)
    norms_list = norms.all()
    return {n.id: n for n in norms_list}


async def _get_users_for_norms(
    db_session: AsyncSession,
    norms_by_id: Dict[int, 'Norms'],
) -> Tuple[Dict[int, List['User']], Dict[int, List['User']]]:
    profession_ids = set()
    activity_ids = set()
    for norm in norms_by_id.values():
        if norm.profession_id:
            profession_ids.add(norm.profession_id)
        if norm.activity_id:
            activity_ids.add(norm.activity_id)

    if not profession_ids and not activity_ids:
        return {}, {}

    query = (
        select(User)
        .outerjoin(ActivityRegistry, ActivityRegistry.user_id == User.id)
        .where(
            or_(
                User.profession_id.in_(profession_ids),
                ActivityRegistry.activity_id.in_(activity_ids),
            )
        )
        .distinct(User.id)
        .options(selectinload(User.activities))
    )

    users = await db_session.scalars(query)
    users_list = users.all()
    prof_users_map = defaultdict(list)
    act_users_map = defaultdict(list)

    for user in users_list:
        if user.profession_id in profession_ids:
            prof_users_map[user.profession_id].append(user)
        for act in user.activities:
            if act.id in activity_ids:
                act_users_map[act.id].append(user)

    return prof_users_map, act_users_map


def _get_need_excel_file(result: dict) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'NeedData'
    ws.cell(row=1, column=1, value=result['size_type'].value)
    ws.cell(row=1, column=2, value='Norm')
    ws.cell(row=1, column=3, value='Given')
    ws.cell(row=1, column=4, value='Need')
    row_idx = 2
    ws.cell(row=row_idx, column=1, value='Total')
    ws.cell(row=row_idx, column=2, value=result['result']['total']['norm'])
    ws.cell(row=row_idx, column=3, value=result['result']['total']['given'])
    ws.cell(row=row_idx, column=4, value=result['result']['total']['need'])
    row_idx = 3
    for size, data in sorted(result['result']['structure'].items()):
        ws.cell(row=row_idx, column=1, value=size)
        ws.cell(row=row_idx, column=2, value=data['norm'])
        ws.cell(row=row_idx, column=3, value=data['given'])
        ws.cell(row=row_idx, column=4, value=data['need'])
        row_idx += 1
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        BytesIO(bio.getvalue()),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="structure.xlsx"'},
    )


def _calculate_material_need(
    material_type: 'MaterialTypes',
    norm_materials: List['NormMaterials'],
    norms_by_id: Dict[int, 'Norms'],
    prof_users_map: Dict[int, List['User']],
    act_users_map: Dict[int, List['User']],
    with_height: bool = False,
) -> Dict[
    str, Union[str, Dict[str, Dict[str, Dict[str, int]]]]
] | StreamingResponse:
    all_users = set()
    size_type_for_material = material_type.size_type or 'no_size_type'

    for nm in norm_materials:
        norm = norms_by_id.get(nm.norm_id)
        if not norm:
            continue

        if norm.profession_id is not None:
            for u in prof_users_map.get(norm.profession_id, []):
                all_users.add(u)
        if norm.activity_id is not None:
            for u in act_users_map.get(norm.activity_id, []):
                all_users.add(u)

    user_norm_sum = defaultdict(int)
    user_given_sum = defaultdict(int)

    for nm in norm_materials:
        norm = norms_by_id.get(nm.norm_id)
        if not norm:
            continue
        users_list = []
        if norm.profession_id is not None:
            users_list.extend(prof_users_map.get(norm.profession_id, []))
        if norm.activity_id is not None:
            users_list.extend(act_users_map.get(norm.activity_id, []))

        for user in set(
            users_list
        ):  # set, чтобы не дублировать, если пересекается
            user_norm_sum[user] += nm.quantity

    for user in all_users:
        total_user_has_for_material = 0
        for um in user.materials:
            if um.material_type_id == material_type.id:
                total_user_has_for_material += um.quantity
        user_given_sum[user] = total_user_has_for_material

    if with_height:
        size_norm = defaultdict(lambda: defaultdict(int))
        size_given = defaultdict(lambda: defaultdict(int))
    else:
        size_norm = defaultdict(int)
        size_given = defaultdict(int)

    for user in all_users:
        if size_type_for_material and size_type_for_material != 'no_size_type':
            user_size = user.additional_features.get(size_type_for_material)
            if user_size is not None:
                size_str = str(user_size)
            else:
                size_str = 'without_size'
        else:
            size_str = 'without_size'

        if with_height:
            user_height = user.additional_features.get('height')
            if user_height is None:
                height_str = 'without_height'
            else:
                height_str = str(user_height)

            size_norm[size_str][height_str] += user_norm_sum[user]
            size_given[size_str][height_str] += user_given_sum[user]
        else:
            size_norm[size_str] += user_norm_sum[user]
            size_given[size_str] += user_given_sum[user]

    structure = {}
    if with_height:
        for size_str in size_norm.keys():
            structure[size_str] = {}
            all_height_keys = set(size_norm[size_str].keys()).union(
                size_given[size_str].keys()
            )
            for h_str in all_height_keys:
                n_val = size_norm[size_str][h_str]
                g_val = size_given[size_str][h_str]
                need_val = max(n_val - g_val, 0)
                structure[size_str][h_str] = {
                    'norm': n_val,
                    'given': g_val,
                    'need': need_val,
                }
        total_norm = 0
        total_given = 0
        total_need = 0
        for size_str, height_map in structure.items():
            for h_str, data in height_map.items():
                total_norm += data['norm']
                total_given += data['given']
                total_need += data['need']
    else:
        all_size_keys = set(size_norm.keys()).union(size_given.keys())
        for size_str in all_size_keys:
            n_val = size_norm[size_str]
            g_val = size_given[size_str]
            need_val = max(n_val - g_val, 0)
            structure[size_str] = {
                'norm': n_val,
                'given': g_val,
                'need': need_val,
            }
        total_norm = sum(s['norm'] for s in structure.values())
        total_given = sum(s['given'] for s in structure.values())
        total_need = sum(s['need'] for s in structure.values())

    result = {
        'size_type': size_type_for_material,
        'result': {
            'total': {
                'norm': total_norm,
                'given': total_given,
                'need': total_need,
            },
            'structure': structure,
        },
    }
    if with_height:
        return result
    return _get_need_excel_file(result)


async def _get_material_type(
    db_session: AsyncSession, material_type_id: int
) -> MaterialTypes:
    query = select(MaterialTypes).where(MaterialTypes.id == material_type_id)
    result = await db_session.scalars(query)
    return result.one_or_none()


async def calculate_need_process(
    db_session: AsyncSession,
    material_type_id: bool = False,
    with_height: bool = False,
) -> Dict[
    str, Union[str, Dict[str, Dict[str, Dict[str, int]]]]
] | StreamingResponse:
    material_type = await _get_material_type(db_session, material_type_id)
    norm_materials = await _get_norm_materials(db_session, material_type_id)
    if not norm_materials:
        return {}
    norms_by_id = await _get_norms_for_materials(db_session, norm_materials)
    prof_users_map, act_users_map = await _get_users_for_norms(
        db_session, norms_by_id
    )
    return _calculate_material_need(
        material_type,
        norm_materials,
        norms_by_id,
        prof_users_map,
        act_users_map,
        with_height=with_height,
    )

async def calculate_need_all_materials_simple(db_session: AsyncSession) -> Dict[str, Dict[str, Union[int, Dict[str, int]]]]:
    material_types_query = select(MaterialTypes)
    material_types = (await db_session.scalars(material_types_query)).all()

    results = {}

    for material_type in material_types:
        norm_materials = await _get_norm_materials(db_session, material_type.id)
        if not norm_materials:
            continue
        norms_by_id = await _get_norms_for_materials(db_session, norm_materials)
        if not norms_by_id:
            continue

        prof_users_map, act_users_map = await _get_users_for_norms(db_session, norms_by_id)

        all_users = set()
        for nm in norm_materials:
            norm = norms_by_id.get(nm.norm_id)
            if not norm:
                continue
            if norm.profession_id is not None:
                all_users.update(prof_users_map.get(norm.profession_id, []))
            if norm.activity_id is not None:
                all_users.update(act_users_map.get(norm.activity_id, []))

        total_norm = 0
        total_given = 0

        for user in all_users:
            user_norm_sum = 0
            for nm in norm_materials:
                norm = norms_by_id.get(nm.norm_id)
                if not norm:
                    continue
                if norm.profession_id == user.profession_id or (
                    norm.activity_id
                    and any(act.id == norm.activity_id for act in user.activities)
                ):
                    user_norm_sum += nm.quantity
            total_norm += user_norm_sum

            user_given_sum = 0
            for um in user.materials:
                if um.material_type_id == material_type.id:
                    user_given_sum += um.quantity
            total_given += user_given_sum

        total_need = max(total_norm - total_given, 0)

        results[material_type.title] = {
            "id": material_type.id,
            "data": {
                "norm": total_norm,
                "given": total_given,
                "need": total_need,
            },
        }

    return results


def _parse_range_str(range_str: str) -> Tuple[int, int]:
    parts = range_str.split('-')
    start = int(parts[0])
    end = int(parts[1])
    return start, end


def _get_table_excel_file(
    table_data: dict,
    extended_size_range: List[str],
    extended_height_range: List[str],
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'NeedData'
    ws.cell(row=1, column=1, value='Size/Height')
    for col_idx, hr in enumerate(extended_height_range, start=2):
        ws.cell(row=1, column=col_idx, value=hr)
    row_idx = 2
    for sr in extended_size_range:
        ws.cell(row=row_idx, column=1, value=sr)
        for col_idx, hr in enumerate(extended_height_range, start=2):
            val = table_data[sr][hr]
            ws.cell(row=row_idx, column=col_idx, value=val)
        row_idx += 1
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        BytesIO(bio.getvalue()),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="needs_table.xlsx"'},
    )


async def calculate_table_process(
    db_session,
    material_type_id: int,
    size_range: List[str],
    height_range: List[str],
    like_file: bool = False,
):
    calc_result = await calculate_need_process(
        db_session=db_session,
        material_type_id=material_type_id,
        with_height=True,
    )

    if not calc_result:
        if like_file:
            return b''
        else:
            return {}

    structure_dict = calc_result['result']['structure']
    extended_size_range = list(size_range) + ['without_size']
    extended_height_range = list(height_range) + ['without_height']

    table_data = {
        sr: {hr: 0 for hr in extended_height_range}
        for sr in extended_size_range
    }

    for size_str, height_map in structure_dict.items():
        if size_str == 'without_size':
            matched_size_range = 'without_size'
        else:
            try:
                s_val = int(size_str)
                matched_size_range = None
                for sr in size_range:
                    start, end = _parse_range_str(sr)
                    if start <= s_val <= end:
                        matched_size_range = sr
                        break
                if matched_size_range is None:
                    matched_size_range = 'without_size'
            except ValueError:
                matched_size_range = 'without_size'

        for h_str, data_dict in height_map.items():
            needed = data_dict['need']
            if h_str == 'without_height':
                matched_height_range = 'without_height'
            else:
                try:
                    h_val = int(h_str)
                    matched_height_range = None
                    for hr in height_range:
                        start, end = _parse_range_str(hr)
                        if start <= h_val <= end:
                            matched_height_range = hr
                            break
                    if matched_height_range is None:
                        matched_height_range = 'without_height'
                except ValueError:
                    matched_height_range = 'without_height'
            table_data[matched_size_range][matched_height_range] += needed

    if like_file:
        return _get_table_excel_file(
            table_data, extended_size_range, extended_height_range
        )
    return table_data
