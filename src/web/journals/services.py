import logging
import os
import uuid
from datetime import datetime
from typing import Sequence

import aiofiles as aiof
from fastapi import UploadFile, Request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy import and_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from database.models import Journals, User, Instructions
from database.orm import Session
from settings import (
    SIGNATURES_DIR,
    BASE_URL,
    STATIC_FOLDER,
    SIGNATURES_FOLDER,
)
from web.exceptions import BulkChecksError
from web.journals.shemas import BulkUpdateJournalsInput

logger = logging.getLogger('control')


def is_valid_uuid(uuids_list: list[str]) -> bool:
    for user_uuid in uuids_list:
        try:
            uuid.UUID(user_uuid, version=4)
        except ValueError:
            return False
    return True


async def check_for_bulk_operation(
    db_session: AsyncSession, bulk_update: BulkUpdateJournalsInput
) -> None:
    if not is_valid_uuid(bulk_update.user_uuids_list):
        raise BulkChecksError('Invalid user_uuids_list')
    len_user_uuids_list = len(bulk_update.user_uuids_list)
    query = select(User.id).where(User.id.in_(bulk_update.user_uuids_list))
    users_ids = (await db_session.scalars(query)).all()
    len_users_ids = len(users_ids)
    if len_user_uuids_list != len_users_ids:
        raise BulkChecksError('Some users not found')
    query = select(Instructions).where(
        Instructions.id == bulk_update.instruction_id
    )
    instruction = await db_session.scalar(query)
    if instruction is None:
        raise BulkChecksError('Instruction not found')


def collect_full_links(journal: Journals, request: Request) -> None:
    from web.instructions.services import get_full_link as get_full_link_ins

    if journal.signature is not None:
        journal.link = get_full_link(request, journal.signature)
    if journal.instruction.filename is not None:
        journal.instruction.link = get_full_link_ins(
            request, journal.instruction.filename
        )
    for history in journal.histories:
        if history.signature is not None:
            history.link = get_full_link(request, history.signature)


async def get_or_create_journals(
    db_session: AsyncSession, instructions_ids: list[int], user_uuid: uuid
):
    query = select(Journals).where(Journals.user_uuid == user_uuid)
    journals = await db_session.scalars(query)
    exist_instructions_ids = []
    exist_journals_ids = []
    list_to_delete = []
    for journal in journals:
        if journal.instruction_id not in instructions_ids:
            list_to_delete.append(journal.id)
        else:
            exist_journals_ids.append(journal.id)
            exist_instructions_ids.append(journal.instruction_id)

    if exist_journals_ids:
        query = (
            update(Journals)
            .where(
                and_(
                    Journals.id.in_(exist_journals_ids),
                    Journals.actual == False,
                )
            )
            .values(actual=True)
        )
        await db_session.execute(query)
    if list_to_delete:
        query = (
            update(Journals)
            .where(Journals.id.in_(list_to_delete))
            .values(actual=False)
        )
        await db_session.execute(query)
    list_to_create = list(set(instructions_ids) - set(exist_instructions_ids))
    for instruction_id in list_to_create:
        journal = Journals(
            instruction_id=instruction_id, user_uuid=user_uuid, actual=True
        )
        db_session.add(journal)
        logger.info(
            f'Created new journal for user {user_uuid} and instruction {instruction_id}'
        )
    await db_session.commit()


async def actualize_journals_for_user(
    user: User,
) -> None:
    if user.profession_id is not None:
        async with Session() as session:
            ins_ids = [
                instruction.id for instruction in user.profession.instructions
            ]
            await get_or_create_journals(session, ins_ids, user.id)
            logger.info(f'Actualized journals for user {user.id}')


def get_full_link(request: Request, filename: str | None) -> str | None:
    if filename is None:
        return None
    base_url = BASE_URL or str(request.base_url)
    return f'{base_url}{STATIC_FOLDER}/{SIGNATURES_FOLDER}/{filename}'


async def add_lines_to_journals_for_new_rule(
    db_session: AsyncSession, profession_id: int, instruction_id: int
) -> None:
    query = select(User.id).where(User.profession_id == profession_id)
    users_ids = await db_session.scalars(query)
    for user_id in users_ids:
        journal = Journals(user_uuid=user_id, instruction_id=instruction_id)
        db_session.add(journal)
    await db_session.commit()


async def remove_lines_to_journals_for_delete_rule(
    db_session: AsyncSession, profession_id: int, instruction_id: int
) -> None:
    query = select(User.id).where(User.profession_id == profession_id)
    users_ids = await db_session.scalars(query)
    query = select(Journals).where(
        and_(
            Journals.user_uuid.in_(users_ids),
            Journals.instruction_id == instruction_id,
        )
    )
    await delete_journals(db_session, query)


async def remove_lines_to_journals_for_delete_ins(
    db_session: AsyncSession, instruction_id: int
) -> None:
    query = select(Journals).where(Journals.instruction_id == instruction_id)
    await delete_journals(db_session, query)


async def remove_lines_to_journals_for_delete_prof(
    db_session: AsyncSession, profession_id: int
) -> None:
    query = select(User.id).where(User.profession_id == profession_id)
    users_ids = await db_session.scalars(query)
    query = select(Journals).where(Journals.user_uuid.in_(users_ids))
    await delete_journals(db_session, query)


async def remove_lines_to_journals_for_delete_user(user: User) -> None:
    async with Session() as db_session:
        query = select(Journals).where(Journals.user_uuid == user.id)
        await delete_journals(db_session, query)
        db_session.commit()


async def delete_journals(db_session: AsyncSession, query) -> None:
    journals = (await db_session.scalars(query)).all()
    # await delete_files_from_journals(journals)
    ids = [journal.id for journal in journals]
    query = update(Journals).where(Journals.id.in_(ids)).values(actual=False)
    await db_session.execute(query)
    logger.info(f'Deleted journals: {ids}')


async def save_file(
    new_file: UploadFile, journal: Journals, user: User
) -> str:
    _, suffix = os.path.splitext(new_file.filename)
    new_name = (
        f'{user.id}--{journal.id}--'
        f"{datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')}{suffix}"
    )
    path_to_file = os.path.join(SIGNATURES_DIR, new_name)
    async with aiof.open(path_to_file, 'wb+') as f:
        await f.write(new_file.file.read())
    logger.info(f'Saved new signature: {new_name}')
    return new_name


async def delete_files_from_journals(journals) -> None:
    for journal in journals:
        if journal.signature:
            await delete_file(journal.signature)
            logger.info(f'Delete signature: {journal.signature}')


async def delete_file(filename: str) -> None:
    try:
        os.remove(os.path.join(SIGNATURES_DIR, filename))
        logger.info(f'Signature {filename} was deleted')
    except FileNotFoundError:
        logger.error(f'Signature {filename} not found for deleting')
        pass


async def get_book(
    instruction: Instructions,
    users: Sequence[User],
    with_history: bool,
    like_file: bool,
    request: Request,
) -> Workbook | list[dict]:
    report_list = []
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A1:Z1')
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    a1 = ws.cell(row=1, column=1)
    a1.value = instruction.title
    a1.font = Font(bold=True)
    a1.alignment = Alignment(horizontal="left")

    ws.append(['Работник', 'Дата ознакомления', 'Подпись электронная', 'Подпись'])

    for user in users:
        user_journal = None
        for journal in user.journals:
            if journal.instruction_id == instruction.id:
                user_journal = journal
                break
        if user_journal is None:
            continue
        if with_history:
            if user_journal is not None:
                if user_journal.histories:
                    for history in user_journal.histories:
                        ws.append(
                            [
                                f'{user.last_name} {user.name} {user.father_name}',
                                str(history.date) if history.date else '',
                                'Да' if history.signature else 'Нет',
                            ]
                        )
                        report_list.append(
                            {
                                'user': f'{user.last_name} {user.name} {user.father_name}',
                                'date': history.date,
                                'signature': get_full_link(request, history.signature),
                            }
                        )
                else:
                    ws.append(
                        [
                            f'{user.last_name} {user.name} {user.father_name}',
                            str(user_journal.last_date_read) if user_journal.last_date_read else '',
                            'Да' if user_journal.signature else 'Нет',
                        ]
                    )
                    report_list.append(
                        {
                            'user': f'{user.last_name} {user.name} {user.father_name}',
                            'date': user_journal.last_date_read,
                            'signature': get_full_link(request, user_journal.signature),
                        }
                    )
        else:
            ws.append(
                [
                    f'{user.last_name} {user.name} {user.father_name}',
                    str(user_journal.last_date_read) if user_journal.last_date_read else '',
                    'Да' if user_journal.signature else 'Нет',
                ]
            )
            report_list.append(
                {
                    'user': f'{user.last_name} {user.name} {user.father_name}',
                    'date': user_journal.last_date_read,
                    'signature': get_full_link(request, user_journal.signature),
                }
            )
    if like_file:
        return wb
    return report_list



